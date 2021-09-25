#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2021/7/28 10:15
# @Author : YXH
# @Email : 874591940@qq.com
# @desc : 手机号码合成神器软件调用

import time
# 窗口操作
import win32con
import win32gui
# 模拟鼠标键盘
from pymouse import PyMouse
from pykeyboard import PyKeyboard
# 剪贴板
import pyperclip
# excel
import xlrd
import openpyxl

#
# import os
# import psutil
# from pynput.keyboard import Listener, Key
# from multiprocessing import Process

# “run”进程对象
process_run = None


class PhoneNumber(object):
    def __init__(self, config):
        """
        初始化数据
        :param config: 配置参数
        """
        # 创建鼠标、键盘对象
        self.my_mouse = PyMouse()
        self.my_keyboard = PyKeyboard()
        # 相关配置
        self.config = config
        # 软件窗口位置
        self.base_x, self.base_y = 0, 0
        # 保存excel数据
        self.excel_data = []
        # 保存数据的excel对象
        self.save_excel = openpyxl.load_workbook(self.config['excel']['save_path'])
        self.save_sheet = self.save_excel[self.save_excel.sheetnames[0]]
        # 当前要写入的行
        self.save_row = 1

    def get_data(self):
        """
        读取excel数据
        :return:
        """
        excel = xlrd.open_workbook(self.config['excel']['data_path'])
        sheet = excel.sheet_by_index(0)
        for i in range(1, sheet.nrows):
            self.excel_data.append({
                'name': sheet.cell_value(i, self.config['excel']['name_col'] - 1),
                'card_id': sheet.cell_value(i, self.config['excel']['card_id_col'] - 1),
            })
        sheet = excel.sheet_by_index(1)
        # 获取窗口名称
        self.config['window_name'] = sheet.cell_value(0, 0)
        # 获取窗口句柄
        hwnd = win32gui.FindWindow(None, self.config['window_name'])
        win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
        # 分别为左上角x、左上角y、右下角x、右下角y的位置
        window_position = win32gui.GetWindowRect(hwnd)
        self.base_x = window_position[0]
        self.base_y = window_position[1]
        # 获取各控件位置
        self.config['position'] = {
            'name_x': int(sheet.cell_value(1, 1) + self.base_x),
            'name_y': int(sheet.cell_value(2, 1) + self.base_y),
            'id_x': int(sheet.cell_value(1, 2) + self.base_x),
            'id_y': int(sheet.cell_value(2, 2) + self.base_y),
            'btn_x': int(sheet.cell_value(1, 3) + self.base_x),
            'btn_y': int(sheet.cell_value(2, 3) + self.base_y),
            'copy1_x': int(sheet.cell_value(1, 4) + self.base_x),
            'copy1_y': int(sheet.cell_value(2, 4) + self.base_y),
            'copy2_x': int(sheet.cell_value(1, 5) + self.base_x),
            'copy2_y': int(sheet.cell_value(2, 5) + self.base_y)
        }

    def send_keys(self, data):
        """
        清空输入框并输入文本
        :param data: 要输入的文本
        :return:
        """
        # 全选
        self.my_keyboard.press_keys([self.my_keyboard.control_key, 'a'])
        time.sleep(1)
        # 清空输入框
        self.my_keyboard.press_key(self.my_keyboard.backspace_key)
        # 将数据复制到剪贴板
        pyperclip.copy(data)
        # 粘贴
        self.my_keyboard.press_keys([self.my_keyboard.control_key, 'v'])

    def find_result(self, name, card_id):
        """
        查询用户信息
        :param name: 姓名
        :param card_id: 身份证号
        :return:
        """
        # 点击“name”输入框并输入name
        self.my_mouse.click(self.config['position']['name_x'], self.config['position']['name_y'])
        time.sleep(1)
        self.send_keys(name)

        # 点击“id”输入框并输入id
        self.my_mouse.click(self.config['position']['id_x'], self.config['position']['id_y'])
        time.sleep(1)
        self.send_keys(card_id)

        # 点击“查询”按钮
        self.my_mouse.click(self.config['position']['btn_x'], self.config['position']['btn_y'])

    def get_result(self):
        """
        在查询结果页面右键，点击复制全部，并返回粘贴板内容
        :return:
        """
        data = []
        while True:
            # “结果”区域右键，2表示右键
            self.my_mouse.click(self.config['position']['copy1_x'], self.config['position']['copy1_y'], 2)
            time.sleep(1)
            # 点击“复制全部”
            self.my_mouse.click(self.config['position']['copy2_x'], self.config['position']['copy2_y'])
            time.sleep(5)
            if '归属地' in pyperclip.paste():
                data = pyperclip.paste().split('\r\n')
                break
        result = []
        # 获取剪贴板内容
        for row in data:
            items = row.split('\t')
            # 若item[0]为英文字母，item[1]中有****，则保存
            if items[0] in ['GS', 'JX', 'ZX', 'JG', 'JD', 'ZW', 'GR', 'RS', 'XY', 'HK', 'JS', 'LN']:
                if '****' in items[1]:
                    result.append(items)
            elif items[0] in ['名下LT数', '备用', '归属地']:
                result.append(items)
        return result

    def write_excel_data(self, data):
        """
        将数据写入excel
        :param data: 要写入的数据
        :return:
        """
        # 将结果依次写入
        for info in data['result']:
            self.save_sheet.cell(self.save_row, 1).value = data['name']
            self.save_sheet.cell(self.save_row, 2).value = data['card_id']
            for index, item in enumerate(info):
                self.save_sheet.cell(self.save_row, index + 3).value = item
            self.save_row += 1
        # 保存excel
        self.save_excel.save(self.config['excel']['save_path'])

    def run(self):
        """
        启动
        :return:
        """
        # 获取excel要查询的数据和窗口位置
        self.get_data()
        print(self.excel_data)
        # 遍历查询
        for info in self.excel_data:
            # 查询用户信息
            self.find_result(info['name'], info['card_id'])
            # 获取并复制查询结果
            info['result'] = self.get_result()
            print(info['result'])
            # 写入excel
            self.write_excel_data(info)
        # 保存excel
        self.save_excel.save(self.config['excel']['save_path'])


def run():
    config = {
        'excel': {
            'data_path': 'info.xlsx',
            'save_path': 'result.xlsx',
            'name_col': 2,
            'card_id_col': 3,
        }
    }
    PhoneNumber(config).run()


if __name__ == '__main__':
    run()
    # pyinstaller -F .\phone_number.py
